import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import os
import sys

class InventoryManagementSystem:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = self.load_or_create_workbook()
        self.sheet = self.wb.active
        self.setup_sheet()

    def load_or_create_workbook(self):
        if os.path.exists(self.file_name):
            return openpyxl.load_workbook(self.file_name)
        wb = Workbook()
        wb.save(self.file_name)
        return wb

    def setup_sheet(self):
        if self.sheet['A1'].value is None:
            headers = ['Product', 'Quantity', 'Price']
            self.sheet.append(headers)
        self.save()

    def save(self):
        self.wb.save(self.file_name)

    def add_product(self, product, quantity, price):
        self.sheet.append([product, quantity, price])
        self.save()

    def update_stock(self, product, quantity_sold):
        for row in self.sheet.iter_rows(min_row=2, max_col=3, max_row=self.sheet.max_row):
            if row[0].value == product:
                current_quantity = row[1].value
                new_quantity = current_quantity - quantity_sold
                if new_quantity < 0:
                    return False, "Insufficient stock"
                row[1].value = new_quantity
                self.save()
                if new_quantity < 5:
                    return True, f"Low stock alert for {product}! Only {new_quantity} left."
                return True, "Stock updated successfully"
        return False, "Product not found"

    def search_product(self, product):
        for row in self.sheet.iter_rows(min_row=2, max_col=3, max_row=self.sheet.max_row):
            if row[0].value == product:
                return f"Product: {row[0].value}, Quantity: {row[1].value}, Price: {row[2].value}"
        return "Product not found"

    def get_all_products(self):
        return [(row[0].value, row[1].value, row[2].value) for row in self.sheet.iter_rows(min_row=2, max_col=3, max_row=self.sheet.max_row)]

class InventoryApp(tk.Tk):
    def __init__(self, inventory_system):
        super().__init__()
        self.inventory_system = inventory_system
        self.title("Inventory Management System")
        self.geometry("800x600")
        self.configure(bg='#f0f0f0')

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.create_widgets()
        self.create_menu()
        self.update_product_list()

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Left Frame for buttons and search
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        ttk.Button(left_frame, text="Add Product", command=self.add_product).pack(fill=tk.X, pady=5)
        ttk.Button(left_frame, text="Sell Product", command=self.sell_product).pack(fill=tk.X, pady=5)

        # Search bar
        ttk.Label(left_frame, text="Search Product:").pack(fill=tk.X, pady=(20, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.update_product_list)
        ttk.Entry(left_frame, textvariable=self.search_var).pack(fill=tk.X)

        # Right Frame for product list
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Product List
        self.tree = ttk.Treeview(right_frame, columns=('Product', 'Quantity', 'Price'), show='headings')
        self.tree.heading('Product', text='Product')
        self.tree.heading('Quantity', text='Quantity')
        self.tree.heading('Price', text='Price')
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

    def create_menu(self):
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Exit", command=self.quit)

    def update_product_list(self, *args):
        self.tree.delete(*self.tree.get_children())
        products = self.inventory_system.get_all_products()
        search_term = self.search_var.get().lower()
        for product in products:
            if search_term in product[0].lower():
                self.tree.insert('', tk.END, values=product)

    def add_product(self):
        product = simpledialog.askstring("Input", "Enter product name:")
        if product:
            quantity = simpledialog.askinteger("Input", "Enter quantity:")
            if quantity is not None:
                price = simpledialog.askfloat("Input", "Enter price:")
                if price is not None:
                    self.inventory_system.add_product(product, quantity, price)
                    messagebox.showinfo("Success", "Product added successfully!")
                    self.update_product_list()

    def sell_product(self):
        product = simpledialog.askstring("Input", "Enter product name:")
        if product:
            quantity_sold = simpledialog.askinteger("Input", "Enter quantity sold:")
            if quantity_sold is not None:
                success, message = self.inventory_system.update_stock(product, quantity_sold)
                if success:
                    messagebox.showinfo("Success", message)
                    self.update_product_list()
                else:
                    messagebox.showerror("Error", message)

def main():
    inventory_system = InventoryManagementSystem("inventory.xlsx")
    app = InventoryApp(inventory_system)
    app.mainloop()

if __name__ == "__main__":
    main()